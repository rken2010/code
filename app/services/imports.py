import io
import json
import re
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

_REQUIRED_FIELDS = {"code", "quantity", "detail", "unit_cost"}

_DEFAULT_MAPPING = {
    "codigo": "code",
    "cantidad": "quantity",
    "detalle": "detail",
    "costo unitario": "unit_cost",
    "cat. prog.": "category_program",
    "unidad de medida": "unit_measure",
    "costo estimado": "estimated_cost",
}


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def parse_mapping(mapping_json: str | None) -> dict[str, str]:
    if not mapping_json:
        return _DEFAULT_MAPPING
    loaded = json.loads(mapping_json)
    return {_normalize(k): v for k, v in loaded.items()}


def parse_excel_rows(file_bytes: bytes, mapping: dict[str, str]) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Archivo Excel vacío"]

    headers = [(_normalize(str(c)) if c is not None else "") for c in rows[0]]
    translated = [mapping.get(h, "") for h in headers]

    if not _REQUIRED_FIELDS.issubset(set(translated)):
        return [], ["Faltan columnas requeridas: code, quantity, detail, unit_cost"]

    items: list[dict[str, Any]] = []
    errors: list[str] = []
    for idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue

        item: dict[str, Any] = {}
        for col_idx, value in enumerate(row):
            field = translated[col_idx] if col_idx < len(translated) else ""
            if not field:
                continue
            item[field] = value

        if not item:
            continue

        try:
            item["code"] = str(item["code"]).strip()
            item["quantity"] = float(item["quantity"])
            item["unit_cost"] = float(item["unit_cost"])
            item["detail"] = str(item["detail"]).strip()
            if not item["code"] or not item["detail"]:
                raise ValueError("code/detail vacíos")
            items.append(item)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Fila {idx}: {exc}")

    return items, errors


def parse_pdf_preview(file_bytes: bytes, filename: str) -> dict[str, Any]:
    reader = PdfReader(io.BytesIO(file_bytes))
    text = "\n".join((p.extract_text() or "") for p in reader.pages[:2])

    lines = [line.strip() for line in text.splitlines() if line.strip()][:10]
    date_match = re.search(r"(\d{2}/\d{2}/\d{4})", text)

    dep_match = None
    for line in lines:
        upper = line.upper()
        if "DEPENDENCIA" in upper or "UNIDAD EJECUTORA" in upper:
            dep_match = line
            break

    return {
        "filename": filename,
        "detected_date": date_match.group(1) if date_match else None,
        "detected_dependency": dep_match,
        "preview_lines": lines,
    }
